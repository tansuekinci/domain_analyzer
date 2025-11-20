import socket
import requests
import whois
import sys
import re
import time
import os
import pandas as pd
from tabulate import tabulate
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Border, Side # Border ve Side eklendi

# --- CONFIGURATION ---
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'
}

# --- HELPER CLASS TO SILENCE NOISY LIBRARIES ---
class SuppressStderr:
    """Context manager to suppress standard error output"""
    def __enter__(self):
        self._original_stderr = sys.stderr
        self._null = open(os.devnull, 'w')
        sys.stderr = self._null

    def __exit__(self, exc_type, exc_val, exc_tb):
        sys.stderr = self._original_stderr
        self._null.close()

def is_valid_ip(text):
    """Regex check for valid IPv4 format"""
    pattern = r"^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$"
    return re.match(pattern, text) is not None

def clean_val(v):
    if v is None: return "-"
    if isinstance(v, list): return ", ".join(str(x) for x in v if x is not None)
    return str(v)

def get_ip_details(ip):
    """Fetches GeoIP and ISP data"""
    try:
        if not is_valid_ip(ip): return None
        time.sleep(1.2) # Rate Limit Protection
        response = requests.get(f"http://ip-api.com/json/{ip}?fields=status,country,city,isp,org,as", timeout=5)
        return response.json()
    except:
        return None

def get_rapiddns_history(domain):
    """Scrapes and parses historical A records from RapidDNS"""
    history_list = []
    try:
        url = f"https://rapiddns.io/s/{domain}"
        response = requests.get(url, headers=HEADERS, timeout=15)
        if response.status_code != 200: return []

        soup = BeautifulSoup(response.text, 'html.parser')
        table = soup.find('table', {'class': 'table'})
        if not table: return []

        rows = table.find_all('tr')[1:]
        for row in rows:
            cols = row.find_all('td')
            found_ip, found_date, found_type = None, "-", "Unknown"
            
            # Smart Parsing
            for col in cols:
                text = col.text.strip()
                if is_valid_ip(text): found_ip = text
                elif re.match(r"\d{4}-\d{2}-\d{2}", text): found_date = text
                elif text in ['A', 'CNAME', 'MX', 'TXT']: found_type = text
            
            if found_type == 'A' and found_ip:
                history_list.append({'ip': found_ip, 'date': found_date})
        return history_list
    except Exception as e:
        return []

def style_history_sheet(ws):
    """Applies bold borders around each record block in the History sheet"""
    thick_side = Side(border_style="thick", color="000000")
    
    # Helper to draw a box around a range of rows
    def draw_box(start_r, end_r):
        # Columns are fixed: A(1) and B(2)
        for r in range(start_r, end_r + 1):
            for c in [1, 2]:
                cell = ws.cell(row=r, column=c)
                
                # Determine borders based on position
                b_top = thick_side if r == start_r else None
                b_bottom = thick_side if r == end_r else None
                b_left = thick_side if c == 1 else None
                b_right = thick_side if c == 2 else None
                
                # Apply border (preserve existing internal structure if needed, but overwrite here is fine)
                # Note: openpyxl requires specifying all sides. 
                # If we want internal lines to be 'none' or 'thin', we leave them default.
                # But we must ensure we don't accidentally remove a side we just set.
                
                # Create a safe existing border check or just build new one
                # Simplest approach: Outline only
                
                current_border = Border(
                    top=b_top if b_top else cell.border.top,
                    bottom=b_bottom if b_bottom else cell.border.bottom,
                    left=b_left if b_left else cell.border.left,
                    right=b_right if b_right else cell.border.right
                )
                cell.border = current_border

    # Iterate to find blocks
    start_row = 2 # Row 1 is header
    max_row = ws.max_row
    
    for row in range(2, max_row + 2): # Go one past end to catch the last block
        cell_val = ws.cell(row=row, column=1).value
        
        # Check if we hit a new record header OR end of sheet
        is_header = cell_val and str(cell_val).startswith("--- RECORD")
        is_end = (row > max_row)
        
        if is_header or is_end:
            # Draw box around the *previous* block
            if row > start_row:
                draw_box(start_row, row - 1)
            
            # Reset start pointer to current row
            start_row = row

def main(url):
    domain = url.replace("http://", "").replace("https://", "").replace("www.", "").split('/')[0]
    print(f"\n🔍 INITIATING DEEP ANALYSIS FOR: {domain} (v9.2 - Boxed Report)\n")

    current_data_list = []
    history_data_list = []
    
    # --- 1. CURRENT STATUS ---
    current_ip = "Unresolved"
    try: current_ip = socket.gethostbyname(domain)
    except: pass

    registrar = "Hidden/Error"
    emails = "-"
    try:
        with SuppressStderr():
            w = whois.whois(domain)
            registrar = clean_val(w.registrar)
            if w.emails:
                emails = w.emails
                if isinstance(emails, list): emails = emails[0]
    except: pass

    geo_current = get_ip_details(current_ip)

    current_data_list.append(["Target Domain", domain])
    current_data_list.append(["Current IP", current_ip])
    
    if geo_current and geo_current.get('status') == 'success':
        current_data_list.append(["Location", f"{geo_current.get('city')} / {geo_current.get('country')}"])
        current_data_list.append(["ISP", f"{geo_current.get('isp')}"])
        current_data_list.append(["ASN", f"{geo_current.get('as')}"])
    
    current_data_list.append(["Registrar", registrar])
    current_data_list.append(["Contact Email", str(emails)])

    # --- 2. HISTORICAL ANALYSIS ---
    print("⏳ Fetching historical records (RapidDNS)...")
    history = get_rapiddns_history(domain)
    
    seen_ips = set()
    seen_ips.add(current_ip)
    found_count = 0
    
    if history:
        print(f"⚡ Found {len(history)} historical records. Analyzing GeoIP data...")
        for i, record in enumerate(history):
            h_ip = record['ip']
            h_date = record['date']
            
            if h_ip in seen_ips: continue
            seen_ips.add(h_ip)
            
            sys.stdout.write(f"\r>> Processing: {h_ip} ({i+1}/{len(history)})   ")
            sys.stdout.flush()
            
            h_geo = get_ip_details(h_ip)
            
            # Data Structure for Excel
            history_data_list.append([f"--- RECORD #{found_count+1} ---", f"Date: {h_date}"])
            history_data_list.append(["Detected IP", h_ip])
            
            if h_geo and h_geo.get('status') == 'success':
                history_data_list.append(["Old Location", f"{h_geo.get('city')} / {h_geo.get('country')}"])
                history_data_list.append(["Old ISP", f"{h_geo.get('isp')}"])
                history_data_list.append(["Old Org.", f"{h_geo.get('org')}"])
            else:
                history_data_list.append(["IP Info", "Location data unavailable"])
            
            found_count += 1
            
        print("\n\n✅ Analysis complete.")
            
    if found_count == 0:
        history_data_list.append(["HISTORY", "No different historical 'A' records found."])

    # --- CONSOLE DISPLAY ---
    combined_view = [["=== CURRENT STATUS ===", "---"]] + current_data_list + \
                    [["=== HISTORICAL DATA ===", "---"]] + history_data_list
    print(tabulate(combined_view, headers=["PARAMETER", "DETAILS"], tablefmt="fancy_grid"))

    # --- EXCEL EXPORT ---
    try:
        ask = input(f"\n💾 Save results to '{domain}.xlsx'? (Y/N): ").lower()
        if ask == 'y' or ask == 'yes':
            filename = f"{domain}.xlsx"
            
            df_current = pd.DataFrame(current_data_list, columns=["PARAMETER", "DETAILS"])
            df_history = pd.DataFrame(history_data_list, columns=["PARAMETER", "DETAILS"])
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df_current.to_excel(writer, sheet_name='Current_Status', index=False)
                df_history.to_excel(writer, sheet_name='History_Logs', index=False)
                
                # Apply Styling
                workbook = writer.book
                
                # --- Sheet 1 Style ---
                ws_curr = writer.sheets['Current_Status']
                for cell in ws_curr[1]: cell.font = Font(bold=True) # Header
                for row in ws_curr.iter_rows(min_row=2, max_col=1): # Col A
                    for cell in row: cell.font = Font(bold=True)
                ws_curr.column_dimensions['A'].width = 30
                ws_curr.column_dimensions['B'].width = 60

                # --- Sheet 2 Style (The Boxed History) ---
                ws_hist = writer.sheets['History_Logs']
                for cell in ws_hist[1]: cell.font = Font(bold=True)
                for row in ws_hist.iter_rows(min_row=2, max_col=1):
                    for cell in row: cell.font = Font(bold=True)
                ws_hist.column_dimensions['A'].width = 30
                ws_hist.column_dimensions['B'].width = 60
                
                # APPLY THE BOX LOGIC
                style_history_sheet(ws_hist)

            print(f"\n[+] SUCCESS: Report saved as '{filename}'")
            print(f"    - Sheet 1: Current_Status")
            print(f"    - Sheet 2: History_Logs (With Bold Boxes)")
        else:
            print("\n[-] Export cancelled.")
    except Exception as e:
        print(f"\n[!] File Error: {e}")

if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else input("Enter target domain (e.g., google.com): ")
    main(target)